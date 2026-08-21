#!/usr/bin/env python3
"""Nạp các file đề thi JLPT (*.xlsx) trong cùng thư mục -> exams-data.js cho quizz.html.

Chạy lại mỗi khi thêm đề mới hoặc sửa file Excel:
    python3 import-exam-xlsx-to-js.py

Hỗ trợ 2 kiểu file (tự nhận dạng, có thể trộn):
  KIỂU A — mỗi phần thi một sheet (vd JLPT_N3_T7_2010_de_thi.xlsx)
      1_文字・語彙 / 2_文法 : 大問 | 問 | 問題文 | 選択肢 1..4 | Đáp án | Ghi chú
      3_読解                : 大問 | 本文 | 問 | 設問 | 選択肢 1..4 | Đáp án
      4_聴解                : 大問 | 番 | メモ / 形式 | 選択肢 1..4 | Đáp án
      5_本文(...)           : セクション | タイトル | 本文     <- đoạn văn đọc hiểu
  KIỂU B — một sheet phẳng cho cả đề (vd JLPT_N3_de_thi_va_dap_an.xlsx)
      STT | Phần thi | 問題 | Câu | Bài đọc (文章) | Câu hỏi | 1 | 2 | 3 | 4 | Đáp án
      (cột 'Bài đọc (文章)' lặp lại cùng một đoạn văn ở mọi câu dùng nó -> tự gom, khử trùng)
      Sheet 'Đáp án ...' (Kỳ thi | Phần thi | 問題 | Câu | Đáp án) -> kho đáp án nhiều kỳ

Các đề cùng một kỳ thi (vd 7/2010) nằm ở nhiều file sẽ được GỘP làm một:
lấy câu hỏi/đáp án từ bản đầy đủ nhất, lấy đoạn văn từ bản có 本文, đáp án
thiếu thì lấy từ kho đáp án.
"""
import glob, json, os, re, sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    sys.exit("Cần openpyxl: pip3 install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))

# nhãn phần thi -> (key, phút cho cả phần, cần file nghe?)   thứ tự = thứ tự hiển thị
SECTION_META = [
    ('文字・語彙', 'moji',    30, False),
    ('文字',       'moji',    30, False),
    ('語彙',       'goi',     30, False),
    ('文法',       'bunpou',  25, False),
    ('読解',       'dokkai',  45, False),
    ('聴解',       'choukai', 40, True),
]
SEC_ORDER = ['moji', 'goi', 'bunpou', 'dokkai', 'choukai']
OPT_A = ['選択肢 1', '選択肢 2', '選択肢 3', '選択肢 4']
OPT_B = ['1', '2', '3', '4']
TEXT_H = ['設問', '問題文', 'メモ / 形式', 'メモ', 'Câu hỏi', '問題']
NO_H = ['問', '番', 'Câu', 'No.']
ANS_H = ['Đáp án', 'Đáp án đúng', '正解']
BODY_H = ['Bài đọc (文章)', 'Bài đọc', '本文', '文章']
# ô lựa chọn chỉ là chỗ trống (đề in không có lựa chọn: câu nghe / câu hình)
PLACEHOLDER = {'', '-', '－', '—', 'ー', '1', '2', '3', '4', '①', '②', '③', '④', 'None'}


KJ = r'一-鿿'
# （　） ＿＿＿ ＿★＿ ___ : chỗ cần điền in trên đề
BLANK_RE = re.compile(r'[（(]\s*[　\s]*[)）]|[＿_]{2,}[★☆]?[＿_]*|[＿_]+[★☆][＿_]*')


def kanji_core(s):
    m = re.match(f'^[{KJ}]+', s)
    return m.group(0) if m else ''


def kana_tail(s):
    m = re.search(r'[぀-ゟ]+$', s)
    return m.group(0) if m else ''


def load_vocab():
    """Đọc file từ vựng (Mimikara...) -> {cách đọc gốc: {phần Hán tự}} để xác định
    chính xác từ được gạch chân trong 文字・語彙 問題1/問題2. Không có file -> {}."""
    idx = {}
    for f in glob.glob(os.path.join(HERE, '*.xlsx')):
        if not re.search(r'mimikara|tu[_ ]?vung|vocab|từ vựng', os.path.basename(f), re.I):
            continue
        try:
            ws = openpyxl.load_workbook(f, data_only=True).active
        except Exception:
            continue
        hdr = headers(ws)
        ki = next((i for k, i in hdr.items() if '漢字' in k or 'Kanji' in k), 1)
        ri = next((i for k, i in hdr.items() if 'ふりがな' in k or 'đọc' in k.lower()), 2)
        for r in ws.iter_rows(min_row=2, values_only=True):
            if ki >= len(r) or ri >= len(r):
                continue
            word, read = txt(r[ki]), txt(r[ri])
            kc, tail = kanji_core(word), kana_tail(word)
            if not kc or not read:
                continue
            core = read[:-len(tail)] if tail and read.endswith(tail) else read
            idx.setdefault(core, set()).add(kc)
    return idx


def mark_word(text, reading, vocab):
    """Tìm vùng Hán tự trong câu có cách đọc = `reading` (dựa từ điển, không đoán)."""
    best = None
    for core, kanjis in vocab.items():
        if not reading.startswith(core):
            continue
        rest = reading[len(core):]
        for kc in kanjis:
            i = text.find(kc)
            if i < 0:
                continue
            j = i + len(kc)
            k = 0
            while k < len(rest) and j + k < len(text) and text[j + k] == rest[k]:
                k += 1
            if k != len(rest):            # đuôi kana trong câu không khớp -> không phải từ này
                continue
            if not best or len(kc) > best[2]:
                best = (i, j + k, len(kc))
    return best[:2] if best else None


def mark_kana(text, option, vocab):
    """問題2: câu in bằng kana, đáp án là Hán tự -> tìm vùng kana tương ứng."""
    kc, tail = kanji_core(option), kana_tail(option)
    if not kc:
        return None
    for core, kanjis in vocab.items():
        if kc in kanjis:
            i = text.find(core + tail)
            if i >= 0:
                return (i, i + len(core) + len(tail))
    return None


def mark_ranges(text, sec_key, mondai, opts, ans, vocab):
    """[[đầu, cuối], ...] vùng cần bôi đậm + loại ('blank' = chỗ điền, 'word' = từ được hỏi)."""
    b = [[m.start(), m.end()] for m in BLANK_RE.finditer(text)]
    if b:
        return b, 'blank'
    if sec_key != 'moji' or mondai not in ('問題1', '問題2') or not vocab:
        return None, None
    real = [o for o in opts if o and o not in PLACEHOLDER]
    if not real:
        return None, None
    # thử đáp án đúng trước, rồi tới các lựa chọn khác (chỉ lựa chọn đúng mới tra ra)
    order = ([real[ans - 1]] if ans and ans <= len(real) else []) + real
    reading_mode = not any(re.search(f'[{KJ}]', o) for o in real)
    for o in order:
        hit = mark_word(text, o, vocab) if reading_mode else mark_kana(text, o, vocab)
        if hit:
            return [list(hit)], 'word'
    return None, None


def txt(v):
    if v is None:
        return ''
    s = str(v).replace('　', ' ').replace('\r\n', '\n').strip()
    return re.sub(r'[ \t]+\n', '\n', s)


def parse_ans(v):
    s = txt(v)
    m = re.search(r'[1-4１-４①-④]', s)
    if not m:
        return None
    ch = m.group(0)
    for base, off in (('1', 0x31), ('１', 0xFF11), ('①', 0x2460)):
        if base <= ch <= chr(ord(base) + 3):
            return ord(ch) - off + 1
    return None


def sec_of(label):
    """'読解 (Đọc hiểu)' / '3_読解' -> ('読解', 'dokkai', 45, False)"""
    name = re.sub(r'^\d+_', '', txt(label))
    for jp, key, mins, audio in SECTION_META:
        if jp in name:
            return jp, key, mins, audio
    return None


def session_of(*texts):
    """'Đề N3 7-2010' / 'kỳ tháng 7/2010' -> '7/2010'"""
    for t in texts:
        m = re.search(r'(?<!\d)(\d{1,2})\s*[-/_.]\s*(20\d\2)', txt(t)) if False else \
            re.search(r'(?<!\d)(\d{1,2})\s*[-/_.]\s*(20\d\d)(?!\d)', txt(t))
        if m:
            return f'{int(m.group(1))}/{m.group(2)}'
        m = re.search(r'T(\d{1,2})[_\s-]*(20\d\d)', txt(t), re.I)
        if m:
            return f'{int(m.group(1))}/{m.group(2)}'
    return None


def headers(ws):
    return {txt(c.value): i for i, c in enumerate(ws[1]) if txt(c.value)}


def pick(hdr, names):
    for n in names:
        if n in hdr:
            return hdr[n]
    return None


def mk_q(sec_key, mondai, no, text, opts, ans, note='', order=0, vocab=None):
    """Chuẩn hoá 1 câu hỏi. Nếu lựa chọn chỉ là chỗ trống -> đánh dấu blank + số lựa chọn."""
    real = [o for o in opts if o and o not in PLACEHOLDER]
    q = {'qid': f'{sec_key}|{mondai}|{no}', 'no': no, 'text': text,
         'opts': opts if len(real) >= 2 else [], 'ans': ans, 'ord': order}
    if note:
        q['note'] = note
    if len(real) < 2:
        q['blank'] = True
        q['optCount'] = 3 if re.search(r'1\s*[〜~-]\s*3', text) else 4
    hi, kind = mark_ranges(text, sec_key, mondai, opts, ans, vocab or {})
    if hi:
        q['hi'] = hi
        q['hik'] = kind
    return q


# ---------------------------------------------------------------- KIỂU A
def read_passages(wb):
    out = []
    for ws in wb.worksheets:
        if '本文' not in re.sub(r'^\d+_', '', ws.title):
            continue
        hdr = headers(ws)
        ci, ti, bi = pick(hdr, ['セクション']), pick(hdr, ['タイトル']), pick(hdr, ['本文'])
        if ti is None or bi is None:
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            if txt(r[bi]):
                out.append({'sec': txt(r[ci]) if ci is not None else '',
                            'title': txt(r[ti]), 'text': txt(r[bi])})
    return out


def match_passage(passages, sec_jp, mondai, ref):
    if ref:
        for i, p in enumerate(passages):
            if p['title'] and p['title'] in ref:
                return i
    tag = f'{sec_jp} {mondai}'
    for i, p in enumerate(passages):
        if p['sec'] and p['sec'].startswith(tag):
            return i
    return None


def parse_layout_a(wb, path, vocab=None):
    """Mỗi sheet là một phần thi. Trả về exam dict hoặc None."""
    passages = read_passages(wb)
    title = source = ''
    notes = []
    for ws in wb.worksheets:
        if 'Hướng dẫn' not in ws.title:
            continue
        for r in ws.iter_rows(values_only=True):
            k, v = txt(r[0]), txt(r[1]) if len(r) > 1 else ''
            if not v:
                continue
            if k.startswith('Đề thi'):
                title = v
            elif k.startswith('Nguồn'):
                source = v
            elif k.startswith('Lưu ý'):
                notes.append(v)

    sections = OrderedDict()
    for ws in wb.worksheets:
        meta = sec_of(ws.title)
        if not meta or '本文' in ws.title:
            continue
        sec_jp, sec_key, mins, audio = meta
        hdr = headers(ws)
        mi, ni, ti = pick(hdr, ['大問']), pick(hdr, NO_H), pick(hdr, TEXT_H)
        oi = [hdr.get(h) for h in OPT_A]
        ai, gi, pi = pick(hdr, ANS_H), pick(hdr, ['Ghi chú']), pick(hdr, ['本文'])
        if mi is None or not any(o is not None for o in oi):
            continue
        cell = lambda r, i: txt(r[i]) if (i is not None and i < len(r)) else ''
        last_mondai, n = '', 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            mondai = cell(r, mi) or last_mondai or '問題1'
            no = cell(r, ni)
            opts = [cell(r, i) for i in oi]
            if not no and len([o for o in opts if o]) < 2:
                continue                       # dòng ghi chú
            last_mondai = mondai
            n += 1
            q = mk_q(sec_key, mondai, no or str(n), cell(r, ti), opts,
                     parse_ans(r[ai]) if ai is not None and ai < len(r) else None,
                     cell(r, gi), n, vocab)
            p = match_passage(passages, sec_jp, mondai, cell(r, pi))
            if p is not None:
                q['p'] = p
            sections.setdefault(sec_key, {'key': sec_key, 'name': sec_jp, 'minutes': mins,
                                          'needAudio': audio, 'qs': []})['qs'].append((mondai, q))
    if not sections:
        return None
    return {'session': session_of(title, os.path.basename(path)), 'title': title, 'source': source,
            'notes': notes, 'passages': passages, 'secs': sections, 'file': os.path.basename(path)}


# ---------------------------------------------------------------- KIỂU B
def number_passages(passages):
    """Cùng một 大問 có nhiều đoạn văn (vd 読解 問題4 có 4 bài) -> thêm số thứ tự vào tiêu đề."""
    cnt = {}
    for p in passages:
        cnt[p['sec']] = cnt.get(p['sec'], 0) + 1
    seen = {}
    for p in passages:
        if cnt[p['sec']] < 2:
            continue
        seen[p['sec']] = seen.get(p['sec'], 0) + 1
        p['title'] = f"{p['sec']} ({seen[p['sec']]})"


def parse_layout_b(ws, wb_notes, path, vocab=None):
    """Một sheet phẳng = một đề (cột 'Phần thi' cho biết phần)."""
    hdr = headers(ws)
    pi, mi, ni, ti = pick(hdr, ['Phần thi']), pick(hdr, ['問題', '大問']), pick(hdr, NO_H), pick(hdr, TEXT_H)
    oi = [hdr.get(h) for h in OPT_B]
    ai, gi = pick(hdr, ANS_H), pick(hdr, ['Ghi chú'])
    bi = pick(hdr, BODY_H)
    if pi is None or mi is None or not all(o is not None for o in oi):
        return None
    cell = lambda r, i: txt(r[i]) if (i is not None and i < len(r)) else ''
    sections, n = OrderedDict(), 0
    passages, pidx = [], {}                        # đoạn văn đã gặp: khoá = nội dung đã bỏ khoảng trắng
    for r in ws.iter_rows(min_row=2, values_only=True):
        meta = sec_of(cell(r, pi))
        if not meta:
            continue
        sec_jp, sec_key, mins, audio = meta
        mondai = cell(r, mi) or '問題1'
        no = cell(r, ni)
        n += 1
        q = mk_q(sec_key, mondai, no or str(n), cell(r, ti), [cell(r, i) for i in oi],
                 parse_ans(r[ai]) if ai is not None and ai < len(r) else None, cell(r, gi), n, vocab)
        body = cell(r, bi)
        if body:
            key = re.sub(r'\s+', '', body)
            if key not in pidx:
                pidx[key] = len(passages)
                passages.append({'sec': f'{sec_jp} {mondai}', 'title': f'{sec_jp} {mondai}',
                                 'text': body})
            q['p'] = pidx[key]
        sections.setdefault(sec_key, {'key': sec_key, 'name': sec_jp, 'minutes': mins,
                                      'needAudio': audio, 'qs': []})['qs'].append((mondai, q))
    if not sections:
        return None
    number_passages(passages)
    return {'session': session_of(ws.title, os.path.basename(path)), 'title': '', 'source': '',
            'notes': wb_notes, 'passages': passages, 'secs': sections, 'file': os.path.basename(path)}


def parse_key_bank(ws):
    """Sheet 'Đáp án ...' -> {kỳ: {(sec_key, mondai, no): đáp án}}"""
    hdr = headers(ws)
    ki, pi, mi, ni, ai = (pick(hdr, ['Kỳ thi', 'Kỳ']), pick(hdr, ['Phần thi']),
                          pick(hdr, ['問題', '大問']), pick(hdr, NO_H), pick(hdr, ANS_H))
    if None in (ki, pi, mi, ni, ai):
        return {}
    bank = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        sess, meta, a = session_of(r[ki]), sec_of(r[pi]), parse_ans(r[ai])
        if not sess or not meta or not a:
            continue
        bank.setdefault(sess, {})[(meta[1], txt(r[mi]), txt(r[ni]))] = a
    return bank


# ---------------------------------------------------------------- gộp & xuất
def read_file(path, vocab=None):
    """Trả về (danh sách exam thô, kho đáp án) của một file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    exams, bank = [], {}
    notes = []
    for ws in wb.worksheets:
        if 'Hướng dẫn' in ws.title:
            for r in ws.iter_rows(values_only=True):
                v = ' '.join(txt(c) for c in r if txt(c))
                if v.startswith('•') or v.startswith('Lưu ý'):
                    notes.append(v.lstrip('• '))
    for ws in wb.worksheets:
        hdr = headers(ws)
        if 'Kỳ thi' in hdr or 'Kỳ' in hdr:
            bank.update(parse_key_bank(ws))
        elif 'Phần thi' in hdr:
            ex = parse_layout_b(ws, notes, path, vocab)
            if ex:
                exams.append(ex)
    if not exams:                                  # thử kiểu A
        ex = parse_layout_a(wb, path, vocab)
        if ex:
            exams.append(ex)
    return exams, bank


def score(q):
    """độ 'đầy đặn' của một câu: dùng để chọn bản tốt hơn khi gộp."""
    return len(q.get('text', '')) + sum(len(o) for o in q.get('opts', [])) + (5 if q.get('ans') else 0)


def merge(raws):
    """Gộp các bản của cùng một kỳ thi thành 1 đề."""
    out = OrderedDict()
    for ex in raws:
        sid = ex['session'] or ex['file']
        if sid not in out:
            ex = json.loads(json.dumps(ex))         # bản sao độc lập
            out[sid] = ex
            continue
        base = out[sid]
        off = len(base['passages'])
        base['passages'] += ex['passages']
        if ex['title'] and not base['title']:
            base['title'] = ex['title']
        if ex['source'] and not base['source']:
            base['source'] = ex['source']
        base['notes'] = list(dict.fromkeys(base['notes'] + ex['notes']))
        base['file'] += ' + ' + ex['file']
        for sk, sec in ex['secs'].items():
            tgt = base['secs'].setdefault(sk, {'key': sk, 'name': sec['name'], 'minutes': sec['minutes'],
                                               'needAudio': sec['needAudio'], 'qs': []})
            idx = {(m, q['qid']): q for m, q in tgt['qs']}
            for mondai, q in sec['qs']:
                cur = idx.get((mondai, q['qid']))
                if q.get('p') is not None:
                    q['p'] += off                   # dịch chỉ số đoạn văn theo bản gộp
                if not cur:
                    tgt['qs'].append((mondai, q))
                    continue
                if score(q) > score(cur):           # bản mới đầy hơn -> thay nội dung
                    keep_p = cur.get('p', q.get('p'))
                    cur.update(q)
                    if keep_p is not None:
                        cur['p'] = keep_p
                else:
                    for k in ('ans', 'p', 'note'):  # bản cũ đầy hơn -> chỉ bù chỗ thiếu
                        if cur.get(k) in (None, '') and q.get(k) not in (None, ''):
                            cur[k] = q[k]
                if q.get('blank') and cur.get('opts'):
                    cur.pop('blank', None); cur.pop('optCount', None)
    return list(out.values())


def finalize(ex, bank):
    """Sắp thứ tự, bù đáp án từ kho, đổi sang cấu trúc mà app dùng."""
    sess = ex['session']
    keys = bank.get(sess, {})
    filled = 0
    sections = []
    for sk in sorted(ex['secs'], key=lambda k: SEC_ORDER.index(k) if k in SEC_ORDER else 99):
        sec = ex['secs'][sk]
        mondais = OrderedDict()
        for mondai, q in sec['qs']:
            mondais.setdefault(mondai, []).append(q)
        mlist = []
        num = lambda s: int(re.sub(r'\D', '', s) or 0)
        for mondai in sorted(mondais, key=num):
            qs = sorted(mondais[mondai], key=lambda q: (num(q['no']) or q.get('ord', 0)))
            for q in qs:
                if not q.get('ans'):
                    a = keys.get((sk, mondai, q['no']))
                    if a:
                        q['ans'] = a
                        filled += 1
                q.pop('ord', None)
            mlist.append({'key': mondai, 'questions': qs})
        total = sum(len(m['questions']) for m in mlist)
        for m in mlist:
            m['minutes'] = max(3, round(sec['minutes'] * len(m['questions']) / total)) if total else sec['minutes']
        sections.append({'key': sk, 'name': sec['name'], 'minutes': sec['minutes'],
                         'needAudio': sec['needAudio'], 'mondais': mlist})
    title = ex['title'] or (f'JLPT N3 — kỳ {sess}' if sess else ex['file'])
    if sess and sess not in title:
        title = f'{title} ({sess})'
    eid = 'n3-' + sess.replace('/', '-') if sess else re.sub(r'\W+', '-', ex['file']).lower()
    return {'id': eid, 'session': sess or '', 'title': title, 'source': ex['source'],
            'notes': ex['notes'], 'files': ex['file'], 'passages': ex['passages'],
            'sections': sections}, filled


def main():
    files = sorted(f for f in glob.glob(os.path.join(HERE, '*.xlsx'))
                   if re.search(r'de[_ ]?thi|dap[_ ]?an|JLPT', os.path.basename(f), re.I))
    if not files:
        sys.exit('Không thấy file đề nào trong ' + HERE)
    vocab = load_vocab()
    if vocab:
        print(f'· từ điển đối chiếu: {len(vocab)} cách đọc (dùng để gạch chân từ được hỏi)')
    raws, bank = [], {}
    for f in files:
        ex, bk = read_file(f, vocab)
        raws += ex
        for k, v in bk.items():
            bank.setdefault(k, {}).update(v)
        print(f'· {os.path.basename(f)}: {len(ex)} đề, {sum(len(v) for v in bk.values())} đáp án'
              f'{" (" + str(len(bk)) + " kỳ)" if bk else ""}')

    exams, report = [], []
    for ex in merge(raws):
        fin, filled = finalize(ex, bank)
        exams.append(fin)
        report.append((fin, filled))
    def sort_key(e):                       # kỳ mới nhất lên đầu
        m = re.match(r'(\d+)/(\d+)', e['session'] or '')
        return (int(m.group(2)), int(m.group(1))) if m else (0, 0)
    exams.sort(key=sort_key, reverse=True)

    have = {e['session'] for e in exams}
    key_only = sorted(((s, len(v)) for s, v in bank.items() if s not in have),
                      key=lambda x: (x[0].split('/')[1], x[0].split('/')[0]), reverse=True)

    out = os.path.join(HERE, 'exams-data.js')
    with open(out, 'w', encoding='utf-8') as fh:
        fh.write('/* TỰ SINH bởi import-exam-xlsx-to-js.py — sửa file .xlsx rồi chạy lại, đừng sửa tay. */\n')
        fh.write('window.EXAMS = ' + json.dumps(exams, ensure_ascii=False, separators=(',', ':')) + ';\n')
        fh.write('window.EXAM_KEYS_ONLY = ' + json.dumps([{'session': s, 'n': n} for s, n in key_only],
                                                         ensure_ascii=False, separators=(',', ':')) + ';\n')
    for fin, filled in report:
        nq = sum(len(m['questions']) for s in fin['sections'] for m in s['mondais'])
        na = sum(1 for s in fin['sections'] for m in s['mondais'] for q in m['questions'] if q['ans'])
        nb = sum(1 for s in fin['sections'] for m in s['mondais'] for q in m['questions'] if q.get('hik') == 'blank')
        nw = sum(1 for s in fin['sections'] for m in s['mondais'] for q in m['questions'] if q.get('hik') == 'word')
        print(f'\n✓ {fin["title"]}  [{fin["files"]}]')
        print(f'  đánh dấu: {nb} câu có chỗ điền · {nw} câu gạch chân từ được hỏi')
        print(f'  {nq} câu · {na} câu có đáp án ({filled} câu lấy từ kho đáp án) · {len(fin["passages"])} đoạn văn')
        for s in fin['sections']:
            print(f'    - {s["name"]}: ' + ' '.join(
                f'{m["key"]}({len(m["questions"])})' for m in s['mondais']))
    print(f'\n→ {out} ({os.path.getsize(out)/1024:.1f} KB) · {len(exams)} đề')
    if key_only:
        print(f'· Kho đáp án còn {len(key_only)} kỳ chưa có file đề: '
              + ', '.join(s for s, _ in key_only[:8]) + ('...' if len(key_only) > 8 else ''))


if __name__ == '__main__':
    main()
